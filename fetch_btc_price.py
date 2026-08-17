"""
Bitcoin Price Hourly Logger
-----------------------------------
Fetches current Bitcoin price (in USD and INR) from CoinGecko's free API
and appends it (with a timestamp) to an Excel file.

This script is designed to be run automatically every hour by a
GitHub Actions workflow (see .github/workflows/btc_price.yml),
so it keeps working even if your PC is off.
"""

import requests
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "bitcoin_price_log.xlsx"
API_URL = "https://api.coingecko.com/api/v3/simple/price"
PARAMS = {
    "ids": "bitcoin",
    "vs_currencies": "usd,inr",
    "include_24hr_change": "true"
}


def fetch_btc_price():
    """Fetch current BTC price in USD and INR from CoinGecko."""
    response = requests.get(API_URL, params=PARAMS, timeout=15)
    response.raise_for_status()
    data = response.json()["bitcoin"]
    return {
        "usd": data["usd"],
        "inr": data["inr"],
        "usd_24h_change": round(data.get("usd_24h_change", 0), 2),
    }


def append_to_excel(price_data):
    """Append a new row with timestamp + price data to the Excel log."""
    now = datetime.now()

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "BTC Price Log"
        # Header row
        ws.append(["Date", "Time", "Price (USD)", "Price (INR)", "24h Change (%)"])

    ws.append([
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        price_data["usd"],
        price_data["inr"],
        price_data["usd_24h_change"],
    ])

    wb.save(EXCEL_FILE)
    print(f"Logged BTC price: ${price_data['usd']} / ₹{price_data['inr']} at {now}")


if __name__ == "__main__":
    price_data = fetch_btc_price()
    append_to_excel(price_data)
